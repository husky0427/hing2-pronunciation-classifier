# -*- coding: utf-8 -*-
"""
Created on Mon Apr 17 09:50:41 2017
combine 2 file with ratio 1~10
@author: husky0427
"""

import os
from combined import combinefiles
from openpyxl import Workbook

# initial
readPath = '../data/combine2file/input'
writePath = '../data/combine2file/output'
writeName = input('Please imput the write file name:')

# user
# read file 2
readNameList = os.listdir(readPath)
for n, i in enumerate(os.listdir(readPath)):
    print(n, ' ', i)
readName1 = readNameList[int(input('Please input file1 number:'))]
# read file 2
readNameList = os.listdir(readPath)
for n, i in enumerate(os.listdir(readPath)):
    print(n, ' ', i)
readName2 = readNameList[int(input('Please input file2 number:'))]


r1 = int(input('ratio1:'))
r2 = int(input('ratio2:'))
list1 = [0,1,2,3,4,5,6,7,8,9,10,r1]
list2 = [10,9,8,7,6,5,4,3,2,1,0,r2]
printList = ['0:10', '1:9', '2:8', '3:7', '4:6', '5:5', '6:4', '7:3', '8:2', '9:1', '10:0', f'{r1}:{r2}']

with open(f'{readPath1}/{readName1}', 'r', encoding='utf8') as fr:
    lines1 = fr.readlines()
with open(f'{readPath2}/{readName2}', 'r', encoding='utf8') as fr2:
    lines2 = fr2.readlines()   

for i, j in zip(list1, list2):
    print(f'正在產生... {i} {j}')
    if i == 0:
        continue    
    elif j == 0:
        continue
    temp = combinefiles([lines1, lines2], [i, j])
    cnt = temp[0]
    noFind = temp[1]
    inappli = temp[2]
    applicRate.append(temp[3])
    prec.append(temp[4])
    outList.append(temp[5])
    spFault.append(temp[6])
    spFault2.append(temp[7])
    
    with open(f'{writePath}\\{fileName}{i}_{j}.txt', 'w', encoding='utf8') as fw:
        for line in outList:
            fw.write(line)
    with open(f'{writePath}\\{fileName}{i}_{j}Rp1.txt', 'w', encoding='utf8') as fw:
        for line in f1:
            fw.write(line)
    with open(f'{writePath}\\{fileName}{i}_{j}Rp2.txt', 'w', encoding='utf8') as fw:
        for line in f2:
            fw.write(line)        
    print('finish!')
    

# 製表
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'combined 2 files'
for n, item in enumerate(printList):
    ws.cell(row=n+2, column=1).value = item
    ws.cell(row=n+2, column=2).value = float(prec[n][0])
    ws.cell(row=n+2, column=3).value = float(prec[n][1])
    ws.cell(row=n+2, column=4).value = float(prec[n][2])
    ws.cell(row=n+2, column=6).value = float(cnt[n]) 
    ws.cell(row=n+2, column=7).value = float(noFind[n])
    ws.cell(row=n+2, column=8).value = float(inappli[n]])
    ws.cell(row=n+2, column=9).value = float(applicRate[n])
    
ws.cell(row=1, column=1).value = 'ratio'
ws.cell(row=1, column=2).value = 'score1'
ws.cell(row=1, column=3).value = 'score2'
ws.cell(row=1, column=4).value = 'score3'
ws.cell(row=1, column=6).value = 'total'
ws.cell(row=1, column=7).value = 'no feature'
ws.cell(row=1, column=8).value = 'inapplicable'
ws.cell(row=1, column=9).value = 'applicability'    

wb.save(filename=f'{writePath}/{writeName}table.xlsx')

    
