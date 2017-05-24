# -*- coding: utf-8 -*-
"""
Created on Thu Apr  6 21:00:27 2017

combine test main function
"""
import os
import testing_combine as testcombine
import testing as test
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook

# initial
cnt = []
noFind = []
correctRate = []
prec = []
correctRate1 = []
prec1 = []
inappli = []
applicRate = []

readPath = '../data/training/output'
writePath = '../data/combined/output'

# user
readNameList = os.listdir(readPath)
for n, i in enumerate(os.listdir(readPath)):
    print(n, ' ', i)
readName1 = readNameList[int(input('Please input file1 number:'))]
readName2 = readNameList[int(input('Please input file2 number:'))]

r1 = int(input('ratio1:'))
r2 = int(input('ratio2:'))
posi1 = int(readName1[:-10].split('_')[2])
posi2 = int(readName2[:-10].split('_')[2])

list1 = [0,1,2,3,4,5,6,7,8,9,10,r1]
list2 = [10,9,8,7,6,5,4,3,2,1,0,r2]
printList = ['0:10', '1:9', '2:8', '3:7', '4:6', '5:5', '6:4', '7:3', '8:2', '9:1', '10:0', f'{r1}:{r2}']


for i, j in zip(list1, list2):
    print(i, j, '正在生產...')
    if i == 0:
        temp = test.testing(readPath, writePath, readName2)
    elif j == 0:
        temp = test.testing(readPath, writePath, readName1)
    else:
        temp = testcombine.testing_combined(readPath, writePath, readName1, readName2, i, j)
    cnt.append(temp[0])
    noFind.append(temp[1])
    inappli.append(temp[2])
    applicRate.append(temp[3])
    prec.append(temp[4])
    prec1.append(temp[5])
    
# 製表
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'combined all'
for n, item in enumerate(printList):
    ws.cell(row=n+2, column=1).value = item
    ws.cell(row=n+2, column=2).value = float(prec[n][0])
    ws.cell(row=n+2, column=3).value = float(prec[n][1])
    ws.cell(row=n+2, column=4).value = float(prec[n][2])
    ws.cell(row=n+2, column=6).value = float(cnt[n])
    ws.cell(row=n+2, column=7).value = float(noFind[n])
    ws.cell(row=n+2, column=8).value = float(inappli[n])
    ws.cell(row=n+2, column=9).value = float(applicRate[n])
ws.cell(row=1, column=1).value = 'ratio'
ws.cell(row=1, column=2).value = 'score1'
ws.cell(row=1, column=3).value = 'score2'
ws.cell(row=1, column=4).value = 'score3'
ws.cell(row=1, column=6).value = 'total'
ws.cell(row=1, column=7).value = 'no feature'
ws.cell(row=1, column=8).value = 'inapplicable'
ws.cell(row=1, column=9).value = 'applicability'

ws1 = wb.create_sheet(title='combined no 1')
for n, item in enumerate(printList):
    ws1.cell(row=n+2, column=1).value = item
    ws1.cell(row=n+2, column=2).value = float(prec1[n][0])
    ws1.cell(row=n+2, column=3).value = float(prec1[n][1])
    ws1.cell(row=n+2, column=4).value = float(prec1[n][2])
ws.cell(row=1, column=1).value = 'ratio'
ws.cell(row=1, column=2).value = 'score1'
ws.cell(row=1, column=3).value = 'score2'
ws.cell(row=1, column=4).value = 'score3'
wb.save(filename=f'{writePath}/{posi1},{posi2}combined.xlsx')
    
    