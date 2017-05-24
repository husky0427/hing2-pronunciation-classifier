# -*- coding: utf-8 -*-
"""
Created on Fri Apr 21 08:00:06 2017
make test excel

@author: husky0427
"""

import glob
import os
from openpyxl import Workbook

readPath = '../data/testing/output'
writePath = '../data/testing/output'
fileList = glob.glob('../data/testing/output/*(test).txt')
readFileName = fileList[0].split('\\')[1]
xlsName = '{0}_{1}'.format(readFileName.split('_')[0], readFileName.split('_')[1])

# 取得檔案內正確率
def getprecision(file):
    with open(file, 'r', encoding='utf8') as f:
        lines = f.readlines()
    total = lines[-8]
    no1 = lines[-7]
    data = lines[-10]
    p = float(total.split('ion:')[1].split('%')[0])
    p1 = float(no1.split('ion:')[1].split('%')[0])
    cnt = float(data.split('共')[1].split('句')[0])
    noFind = float(data.split('feature:')[1].split(', in')[0])
    inappli = float(data.split('ble:')[1].split(', app')[0])
    applicRate = float(data.split('lity:')[1].split('%')[0])
    return cnt, noFind, inappli, applicRate, p, p1

mns1 = getprecision(fileList[0])  # -1
mns2 = getprecision(fileList[1])  # -2
pls1 = getprecision(fileList[2])  # 1
pls2 = getprecision(fileList[3])  # 2


wb = Workbook()

ws = wb.get_active_sheet()
ws.title = 'all'
ws.cell(row=1, column=1).value = 'position'
ws.cell(row=1, column=2).value = 'total'
ws.cell(row=1, column=3).value = 'no feature'
ws.cell(row=1, column=4).value = 'inapplicable'
ws.cell(row=1, column=5).value = 'applicability'
ws.cell(row=1, column=6).value = 'precision(all)'
ws.cell(row=1, column=7).value = 'precision(no 1)'
ws.cell(row=2, column=1).value = -2
ws.cell(row=3, column=1).value = -1
ws.cell(row=4, column=1).value = 1
ws.cell(row=5, column=1).value = 2
for i in range(6):
    ws.cell(row=2, column=i+2).value = mns2[i]
    ws.cell(row=3, column=i+2).value = mns1[i]
    ws.cell(row=4, column=i+2).value = pls1[i]
    ws.cell(row=5, column=i+2).value = pls2[i]

wb.save(filename=f'{writePath}/{xlsName}.xlsx')

print('finish!!')
