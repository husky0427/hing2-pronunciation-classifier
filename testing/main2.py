# -*- coding: utf-8 -*-
"""
Created on Mon Apr 17 09:50:41 2017
combine 2 file
@author: husky0427
"""

import os
import testing_combine as testcombine
import testing as test
from combinefile import *
from openpyxl import Workbook

# initial
readPath1 = '../data/combined/output'
readPath2 = '../data/testing/output'
readPath3 = '../data/training/output'
writePath = '../data/combinefile/output'
_1List = []
correctRate = []
correctRate1 = []
prec = []
prec1 = []

# user
readNameList = os.listdir(readPath1)
for n, i in enumerate(os.listdir(readPath1)):
    print(n, ' ', i)
readName1 = readNameList[int(input('Please input file1 number:'))]

readNameList = os.listdir(readPath2)
for n, i in enumerate(os.listdir(readPath2)):
    print(n, ' ', i)
readName2 = readNameList[int(input('Please input file2 number:'))]

readNameList = os.listdir(readPath3)
keyin = ''
readName3List = []
while keyin != 'x':
    for n, i in enumerate(readNameList):
        print(n, ' ', i)
    keyin = input('Please input trained file1 number:')
    try:
        readName3 = readNameList[int(keyin)]
        readName3List.append(readName3)
        print(readPath3, readName3)
        with open(f'{readPath3}\\{readName3}', 'r') as fr:
            words = fr.readlines()
        for word in words:
            index = word.split()[0]
            freq = int(word.split()[1])
            if freq == 1:
                _1List.append(index)
    except:
        pass

r1 = int(input('ratio1:'))
r2 = int(input('ratio2:'))
list1 = [0,1,2,3,4,5,6,7,8,9,10,r1]
list2 = [10,9,8,7,6,5,4,3,2,1,0,r2]
printList = ['0:10', '1:9', '2:8', '3:7', '4:6', '5:5', '6:4', '7:3', '8:2', '9:1', '10:0', f'{r1}:{r2}']

posi1 = readName1.split('_')[-1][:-4].split('-')
posi2 = int(readName2[:-10].split('_')[2])

with open(f'{readPath1}/{readName1}', 'r', encoding='utf8') as fr:
    lines1 = fr.readlines()
lines1 = lines1[:-12]
with open(f'{readPath2}/{readName2}', 'r', encoding='utf8') as fr2:
    lines2 = fr2.readlines()   
lines2 = lines2[:-12]



for i, j in zip(list1, list2):
    print(f'正在產生... {i} {j}')
    if i == 0:
        temp = test.testing(readPath3, writePath, readName2.replace('test', 'tone'))
        correctRate.append(temp[2])
        prec.append(temp[3])
        correctRate1.append(temp[4])
        prec1.append(temp[5])
    elif j == 0:
        temp = testcombine.testing_combined(readPath3, writePath, readName3List[0], readName3List[1], int(posi1[0]), int(posi1[1]))
        correctRate.append(temp[2])
        prec.append(temp[3])
        correctRate1.append(temp[4])
        prec1.append(temp[5])
    else:
        temp = combinefile(lines1, lines2, i, j, _1List)
        outList = temp[0]
        f1 = temp[1]  # spFault
        f2 = temp[2]  # spFault2
        correctRate.append(temp[3])
        prec.append(temp[4])
        correctRate1.append(temp[5])
        prec1.append(temp[6])
    
        with open(f'{writePath}\\{i}_{j}.txt', 'w', encoding='utf8') as fw:
            for line in outList:
                fw.write(line)
        with open(f'{writePath}\\{i}_{j}Rp1.txt', 'w', encoding='utf8') as fw:
            for line in f1:
                fw.write(line)
        with open(f'{writePath}\\{i}_{j}Rp2.txt', 'w', encoding='utf8') as fw:
            for line in f2:
                fw.write(line)        
    print('finish!')
    

# 製表
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'combined all'
for n, item in enumerate(printList):
    ws.cell(row=n+1, column=1).value = item
    ws.cell(row=n+1, column=2).value = float(correctRate[n][0])
    ws.cell(row=n+1, column=3).value = float(correctRate[n][1])
    ws.cell(row=n+1, column=4).value = float(correctRate[n][2])
    ws.cell(row=n+1, column=5).value = ' '
    ws.cell(row=n+1, column=6).value = float(prec[n][0])
    ws.cell(row=n+1, column=7).value = float(prec[n][1])
    ws.cell(row=n+1, column=8).value = float(prec[n][2])
ws1 = wb.create_sheet(title='combined no 1')
for n, item in enumerate(printList):
    ws1.cell(row=n+1, column=1).value = item
    ws1.cell(row=n+1, column=2).value = float(correctRate1[n][0])
    ws1.cell(row=n+1, column=3).value = float(correctRate1[n][1])
    ws1.cell(row=n+1, column=4).value = float(correctRate1[n][2])
    ws1.cell(row=n+1, column=5).value = ' '
    ws1.cell(row=n+1, column=6).value = float(prec1[n][0])
    ws1.cell(row=n+1, column=7).value = float(prec1[n][1])
    ws1.cell(row=n+1, column=8).value = float(prec1[n][2])
wb.save(filename=f'{writePath}/combinedfile.xlsx')

    