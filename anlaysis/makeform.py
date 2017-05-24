# -*- coding: utf-8 -*-
"""
Created on Wed Apr 26 08:35:46 2017

把training data 生成分析用表
@author: speechlab
"""

import os
from openpyxl import Workbook

# user
readPath = '../data/training/output'
writePath = '../data/analysis/output'
readNameList = os.listdir(readPath)
for n, i in enumerate(os.listdir(readPath), 1):
    print(n, ' ', i)
readName = readNameList[int(input('Please input file1 number:'))-1]

# load file    
with open(f'{readPath}\\{readName}') as fr:
    lines = fr.readlines()

def processline(string):
    word = string.split('[')[0].split()[0]
    freq = int(string.split('[')[0].split()[1])
    temp = string.split('[')[1].split(']')[0]
    score = [int(i) for i in temp.split(', ')]
    return word, freq, score

# open workbook
wb = Workbook()
ws = wb.get_active_sheet()
ws.cell(row=1, column=1).value = 'Word'
ws.cell(row=1, column=2).value = 'Frequency'
ws.cell(row=1, column=3).value = 'do2'
ws.cell(row=1, column=4).value = 'e3-sai4'
ws.cell(row=1, column=5).value = 'ggau2'
ws.cell(row=1, column=6).value = 'giann2'
ws.cell(row=1, column=7).value = 'hang2'
ws.cell(row=1, column=8).value = 'hing2'
ws.cell(row=1, column=9).value = 'hing7'
ws.cell(row=1, column=10).value = 'zua7'

for n, line in enumerate(lines, 2):
    word, freq, score = processline(line)
    ws.cell(row=n, column=1).value = word
    ws.cell(row=n, column=2).value = freq
    ws.cell(row=n, column=3).value = score[0]
    ws.cell(row=n, column=4).value = score[1]
    ws.cell(row=n, column=5).value = score[2]
    ws.cell(row=n, column=6).value = score[3]
    ws.cell(row=n, column=7).value = score[4]
    ws.cell(row=n, column=8).value = score[5]
    ws.cell(row=n, column=9).value = score[6]
    ws.cell(row=n, column=10).value = score[7]
wb.save(filename=f'{writePath}/{readName[:-4]}.xlsx')
print('finish!!')
    
